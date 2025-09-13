import pandas as pd
import numpy as np
from typing import Dict, List, Optional, Union
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.graphics.shapes import Drawing, Line
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics.charts.piecharts import Pie
import os
import matplotlib.pyplot as plt
import seaborn as sns
from jinja2 import Environment, FileSystemLoader
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import plotly.express as px
from loguru import logger


class RiskExporter:
    """风险报告导出器"""

    def __init__(self, output_dir: str = "reports"):
        """
        初始化风险报告导出器

        Args:
            output_dir: 输出目录路径
        """
        self.output_dir = output_dir
        self.template_dir = os.path.join(
            os.path.dirname(__file__), "templates")
        self._setup_logging()

    def _setup_logging(self):
        """设置日志"""
        # Loguru配置在core.loguru_config中统一管理
        self.logger = logger

    def export_report(self, risk_report: Dict, format: str = "excel") -> str:
        """
        导出风险报告

        Args:
            risk_report: 风险报告数据
            format: 导出格式，支持 "excel", "pdf", "csv", "html"

        Returns:
            str: 导出文件路径
        """
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"risk_report_{timestamp}"

            if format.lower() == "excel":
                return self._export_to_excel(risk_report, filename)
            elif format.lower() == "pdf":
                return self._export_to_pdf(risk_report, filename)
            elif format.lower() == "csv":
                return self._export_to_csv(risk_report, filename)
            elif format.lower() == "html":
                return self._export_to_html(risk_report, filename)
            else:
                raise ValueError(f"不支持的导出格式: {format}")

        except Exception as e:
            self.logger.error(f"导出风险报告失败: {str(e)}")
            raise

    def _export_to_excel(self, risk_report: Dict, filename: str) -> str:
        """导出为Excel格式"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows

            output_path = os.path.join(self.output_dir, f"{filename}.xlsx")
            os.makedirs(self.output_dir, exist_ok=True)

            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # 删除默认工作表

            # 创建风险概览工作表
            ws_overview = wb.create_sheet("风险概览")

            # 添加标题
            ws_overview['A1'] = "FactorWeave-Quant 风险报告"
            ws_overview['A1'].font = Font(size=16, bold=True)
            ws_overview['A2'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

            # 风险指标数据
            risk_data = [
                ["指标名称", "当前值", "风险等级", "阈值", "状态"],
                ["VaR(95%)", f"{risk_report.get('var_95', 0):.2f}%", "中等", "5.0%", "正常"],
                ["最大回撤", f"{risk_report.get('max_drawdown', 0):.2f}%", "中等", "15.0%", "正常"],
                ["波动率", f"{risk_report.get('volatility', 0):.2f}%", "低", "20.0%", "正常"],
                ["夏普比率", f"{risk_report.get('sharpe_ratio', 0):.2f}", "良好", "1.0", "优秀"],
                ["Beta系数", f"{risk_report.get('beta', 1):.2f}", "中性", "1.0", "正常"],
                ["仓位风险", f"{risk_report.get('position_risk', 0):.2f}%", "中等", "70.0%", "正常"],
                ["市场风险", f"{risk_report.get('market_risk', 0):.2f}%", "中等", "60.0%", "正常"],
                ["流动性风险", f"{risk_report.get('liquidity_risk', 0):.2f}%", "低", "40.0%", "正常"]
            ]

            # 添加数据到工作表
            for i, row in enumerate(risk_data, start=4):
                for j, value in enumerate(row, start=1):
                    cell = ws_overview.cell(row=i, column=j, value=value)
                    if i == 4:  # 表头
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center")

            # 设置列宽
            ws_overview.column_dimensions['A'].width = 15
            ws_overview.column_dimensions['B'].width = 12
            ws_overview.column_dimensions['C'].width = 12
            ws_overview.column_dimensions['D'].width = 10
            ws_overview.column_dimensions['E'].width = 10

            # 创建历史数据工作表（如果有历史数据）
            if 'historical_data' in risk_report:
                ws_history = wb.create_sheet("历史趋势")
                history_df = pd.DataFrame(risk_report['historical_data'])

                # 添加表头
                for c_idx, column in enumerate(history_df.columns, 1):
                    cell = ws_history.cell(row=1, column=c_idx, value=column)
                    cell.font = Font(bold=True)

                # 添加数据
                for r_idx, row in enumerate(history_df.values, 2):
                    for c_idx, value in enumerate(row, 1):
                        ws_history.cell(row=r_idx, column=c_idx, value=value)

            # 创建告警记录工作表（如果有告警数据）
            if 'alerts' in risk_report:
                ws_alerts = wb.create_sheet("告警记录")
                alerts_data = [
                    ["时间", "规则名称", "风险类型", "告警级别", "消息"],
                ]

                for alert in risk_report['alerts']:
                    alerts_data.append([
                        alert.get('created_at', ''),
                        alert.get('rule_name', ''),
                        alert.get('metric_name', ''),
                        alert.get('alert_level', ''),
                        alert.get('message', '')
                    ])

                for i, row in enumerate(alerts_data, start=1):
                    for j, value in enumerate(row, start=1):
                        cell = ws_alerts.cell(row=i, column=j, value=value)
                        if i == 1:  # 表头
                            cell.font = Font(bold=True)

            # 保存文件
            wb.save(output_path)
            logger.info(f"风险报告已导出到Excel: {output_path}")
            return output_path

        except Exception as e:
            logger.error(f"导出Excel失败: {e}")
            raise

    def _export_to_pdf(self, risk_report: Dict, filename: str) -> str:
        """导出为PDF格式"""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib import colors
            from reportlab.lib.units import inch

            output_path = os.path.join(self.output_dir, f"{filename}.pdf")
            os.makedirs(self.output_dir, exist_ok=True)

            doc = SimpleDocTemplate(output_path, pagesize=A4)
            styles = getSampleStyleSheet()
            story = []

            # 标题
            title = Paragraph("FactorWeave-Quant 风险报告", styles['Title'])
            story.append(title)
            story.append(Spacer(1, 12))

            # 生成时间
            timestamp = Paragraph(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal'])
            story.append(timestamp)
            story.append(Spacer(1, 12))

            # 风险概览表格
            risk_data = [
                ["指标名称", "当前值", "风险等级", "阈值", "状态"],
                ["VaR(95%)", f"{risk_report.get('var_95', 0):.2f}%", "中等", "5.0%", "正常"],
                ["最大回撤", f"{risk_report.get('max_drawdown', 0):.2f}%", "中等", "15.0%", "正常"],
                ["波动率", f"{risk_report.get('volatility', 0):.2f}%", "低", "20.0%", "正常"],
                ["夏普比率", f"{risk_report.get('sharpe_ratio', 0):.2f}", "良好", "1.0", "优秀"],
                ["Beta系数", f"{risk_report.get('beta', 1):.2f}", "中性", "1.0", "正常"]
            ]

            table = Table(risk_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))

            story.append(table)
            story.append(Spacer(1, 12))

            # 添加告警信息（如果有）
            if 'alerts' in risk_report and risk_report['alerts']:
                alert_title = Paragraph("风险告警", styles['Heading2'])
                story.append(alert_title)
                story.append(Spacer(1, 6))

                for alert in risk_report['alerts'][:5]:  # 只显示前5个告警
                    alert_text = f"• {alert.get('message', '')}"
                    alert_para = Paragraph(alert_text, styles['Normal'])
                    story.append(alert_para)

            # 构建PDF
            doc.build(story)
            logger.info(f"风险报告已导出到PDF: {output_path}")
            return output_path

        except Exception as e:
            logger.error(f"导出PDF失败: {e}")
            raise

    def _export_to_csv(self, risk_report: Dict, filename: str) -> str:
        """导出为CSV格式"""
        try:
            output_path = os.path.join(self.output_dir, f"{filename}.csv")
            os.makedirs(self.output_dir, exist_ok=True)

            # 创建风险指标DataFrame
            risk_data = {
                '指标名称': ['VaR(95%)', '最大回撤', '波动率', '夏普比率', 'Beta系数', '仓位风险', '市场风险', '流动性风险'],
                '当前值': [
                    f"{risk_report.get('var_95', 0):.2f}%",
                    f"{risk_report.get('max_drawdown', 0):.2f}%",
                    f"{risk_report.get('volatility', 0):.2f}%",
                    f"{risk_report.get('sharpe_ratio', 0):.2f}",
                    f"{risk_report.get('beta', 1):.2f}",
                    f"{risk_report.get('position_risk', 0):.2f}%",
                    f"{risk_report.get('market_risk', 0):.2f}%",
                    f"{risk_report.get('liquidity_risk', 0):.2f}%"
                ],
                '风险等级': ['中等', '中等', '低', '良好', '中性', '中等', '中等', '低'],
                '阈值': ['5.0%', '15.0%', '20.0%', '1.0', '1.0', '70.0%', '60.0%', '40.0%'],
                '状态': ['正常', '正常', '正常', '优秀', '正常', '正常', '正常', '正常']
            }

            df = pd.DataFrame(risk_data)
            df.to_csv(output_path, index=False, encoding='utf-8-sig')

            logger.info(f"风险报告已导出到CSV: {output_path}")
            return output_path

        except Exception as e:
            logger.error(f"导出CSV失败: {e}")
            raise

    def _export_to_html(self, risk_report: Dict, filename: str) -> str:
        """导出为HTML格式"""
        try:
            output_path = os.path.join(self.output_dir, f"{filename}.html")
            os.makedirs(self.output_dir, exist_ok=True)

            html_content = f"""
            <!DOCTYPE html>
            <html lang="zh-CN">
            <head>
                <meta charset="UTF-8">
                <meta name="viewport" content="width=device-width, initial-scale=1.0">
                <title>FactorWeave-Quant 风险报告</title>
                <style>
                    body {{ font-family: 'Microsoft YaHei', Arial, sans-serif; margin: 20px; }}
                    .header {{ background: #2c3e50; color: white; padding: 20px; border-radius: 5px; margin-bottom: 20px; }}
                    .header h1 {{ margin: 0; }}
                    .timestamp {{ color: #bdc3c7; margin-top: 10px; }}
                    table {{ width: 100%; border-collapse: collapse; margin-bottom: 20px; }}
                    th, td {{ border: 1px solid #ddd; padding: 12px; text-align: left; }}
                    th {{ background-color: #34495e; color: white; }}
                    tr:nth-child(even) {{ background-color: #f2f2f2; }}
                    .status-normal {{ color: #27ae60; font-weight: bold; }}
                    .status-excellent {{ color: #2ecc71; font-weight: bold; }}
                    .alert-section {{ background: #fff3cd; border: 1px solid #ffeaa7; padding: 15px; border-radius: 5px; margin-top: 20px; }}
                    .alert-item {{ margin: 5px 0; }}
                </style>
            </head>
            <body>
                <div class="header">
                    <h1>FactorWeave-Quant 风险报告</h1>
                    <div class="timestamp">生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</div>
                </div>
                
                <h2>风险指标概览</h2>
                <table>
                    <thead>
                        <tr>
                            <th>指标名称</th>
                            <th>当前值</th>
                            <th>风险等级</th>
                            <th>阈值</th>
                            <th>状态</th>
                        </tr>
                    </thead>
                    <tbody>
                        <tr>
                            <td>VaR(95%)</td>
                            <td>{risk_report.get('var_95', 0):.2f}%</td>
                            <td>中等</td>
                            <td>5.0%</td>
                            <td><span class="status-normal">正常</span></td>
                        </tr>
                        <tr>
                            <td>最大回撤</td>
                            <td>{risk_report.get('max_drawdown', 0):.2f}%</td>
                            <td>中等</td>
                            <td>15.0%</td>
                            <td><span class="status-normal">正常</span></td>
                        </tr>
                        <tr>
                            <td>波动率</td>
                            <td>{risk_report.get('volatility', 0):.2f}%</td>
                            <td>低</td>
                            <td>20.0%</td>
                            <td><span class="status-normal">正常</span></td>
                        </tr>
                        <tr>
                            <td>夏普比率</td>
                            <td>{risk_report.get('sharpe_ratio', 0):.2f}</td>
                            <td>良好</td>
                            <td>1.0</td>
                            <td><span class="status-excellent">优秀</span></td>
                        </tr>
                        <tr>
                            <td>Beta系数</td>
                            <td>{risk_report.get('beta', 1):.2f}</td>
                            <td>中性</td>
                            <td>1.0</td>
                            <td><span class="status-normal">正常</span></td>
                        </tr>
                    </tbody>
                </table>
            """

            # 添加告警信息（如果有）
            if 'alerts' in risk_report and risk_report['alerts']:
                html_content += """
                <div class="alert-section">
                    <h3>风险告警</h3>
                """
                for alert in risk_report['alerts']:
                    html_content += f'<div class="alert-item">• {alert.get("message", "")}</div>'
                html_content += "</div>"

            html_content += """
            </body>
            </html>
            """

            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(html_content)

            logger.info(f"风险报告已导出到HTML: {output_path}")
            return output_path

        except Exception as e:
            logger.error(f"导出HTML失败: {e}")
            raise

    def _generate_html_charts(self, risk_report: Dict) -> List[str]:
        """
        生成HTML图表

        Args:
            risk_report: 风险报告数据

        Returns:
            List[str]: 图表HTML代码列表
        """
        charts = []
        try:
            # 生成风险指标图表
            charts.append(self._create_risk_indicators_chart(risk_report))

            # 生成市场风险图表
            charts.append(self._create_market_risk_chart(risk_report))

            # 生成行业风险图表
            charts.append(self._create_sector_risk_chart(risk_report))

            # 生成流动性风险图表
            charts.append(self._create_liquidity_risk_chart(risk_report))

            # 生成组合风险图表
            charts.append(self._create_portfolio_risk_chart(risk_report))

            # 生成预警分析图表
            charts.append(self._create_alert_analysis_chart(risk_report))

        except Exception as e:
            self.logger.error(f"生成HTML图表失败: {str(e)}")

        return charts

    def _create_risk_indicators_chart(self, risk_report: Dict) -> str:
        """创建风险指标图表"""
        # TODO: 实现风险指标图表生成逻辑
        pass

    def _create_market_risk_chart(self, risk_report: Dict) -> str:
        """创建市场风险图表"""
        # TODO: 实现市场风险图表生成逻辑
        pass

    def _create_sector_risk_chart(self, risk_report: Dict) -> str:
        """创建行业风险图表"""
        # TODO: 实现行业风险图表生成逻辑
        pass

    def _create_liquidity_risk_chart(self, risk_report: Dict) -> str:
        """创建流动性风险图表"""
        # TODO: 实现流动性风险图表生成逻辑
        pass

    def _create_portfolio_risk_chart(self, risk_report: Dict) -> str:
        """创建组合风险图表"""
        # TODO: 实现组合风险图表生成逻辑
        pass

    def _create_alert_analysis_chart(self, risk_report: Dict) -> str:
        """创建预警分析图表"""
        # TODO: 实现预警分析图表生成逻辑
        pass

    def export_quality_report(self, quality_report: Dict, format: str = "excel") -> str:
        """
        导出数据质量报告
        Args:
            quality_report: 数据质量报告字典
            format: 导出格式，支持 "excel", "csv", "html"
        Returns:
            str: 导出文件路径
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"quality_report_{timestamp}"
        output_path = os.path.join(
            self.output_dir, f"{filename}.{format if format != 'excel' else 'xlsx'}")
        os.makedirs(self.output_dir, exist_ok=True)
        # 主体表格
        main_df = pd.DataFrame({
            '字段': list(quality_report['field_distributions'].keys()),
            '分布': [str(quality_report['field_distributions'][k]) for k in quality_report['field_distributions'].keys()]
        })
        # 其它统计
        meta = pd.DataFrame({
            '项目': ['context', 'shape', 'columns', 'missing_fields', 'empty_ratio', 'anomaly_stats', 'type_errors', 'price_relation_errors', 'logic_errors', 'quality_score'],
            '内容': [str(quality_report.get(k, '')) for k in ['context', 'shape', 'columns', 'missing_fields', 'empty_ratio', 'anomaly_stats', 'type_errors', 'price_relation_errors', 'logic_errors', 'quality_score']]
        })
        if format == "excel":
            with pd.ExcelWriter(output_path) as writer:
                main_df.to_excel(writer, sheet_name="字段分布", index=False)
                meta.to_excel(writer, sheet_name="报告摘要", index=False)
        elif format == "csv":
            main_df.to_csv(output_path, index=False)
        elif format == "html":
            html = main_df.to_html(index=False) + meta.to_html(index=False)
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(html)
        else:
            raise ValueError(f"不支持的导出格式: {format}")
        return output_path
